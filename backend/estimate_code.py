"""Evidence-led story estimation on Devvy's one shared local model runtime.

Division of labour, and the reason for it:

* ``estimation_framework`` owns every number. Base sum, adjustments, Fibonacci band,
  maturity cap, gates, confidence, recommendation — all plain Python.
* This module owns the *conversation* with the model: what context it sees, what
  contract it must answer in, and how a bad answer is repaired.

The model is asked for one thing only — a 1-5 score and a short reason per factor. It is
never trusted with arithmetic or with delivery policy. When it declines to score a factor
the application falls back to a keyword heuristic and labels that factor ``heuristic`` in
the output, so a reader can always tell judgement apart from a guess.
"""

from __future__ import annotations

import csv
import io
import json
import re
from collections.abc import Callable
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Literal

import httpx
from pydantic import BaseModel, Field, field_validator, model_validator

from backend.config import Settings
from backend.estimation_framework import (
    EFFORT_DEFAULTS,
    FACTOR_BY_ID,
    FACTOR_IDS,
    FACTORS,
    FIBONACCI_POINTS,
    FRAMEWORK_DOCUMENT,
    FRAMEWORK_VERSION,
    MATURITY_TAXONOMY,
    STACK_LABELS,
    Calculation,
    FactorScore,
    StackProfile,
    calculate,
    confidence,
    detailed_reasoning,
    decide,
    estimation_suggestions,
    policy_checks,
    risk_flags,
    spike_template,
)
from backend.harness import ContextSource, assemble_context, GROUNDING_CONTRACT_BRIEF
from backend.eagle import (
    EAGLE_VERSION,
    adversarial_review,
    aggregate,
    attribute_failure,
    build_blackboard,
    build_contract,
    build_snapshot,
    compare_references,
    critic_review,
    debate,
    optimistic_review,
    spike_gate,
    validate,
)
from backend.estimation_pipeline import (
    PROTECTED_FACTORS,
    PIPELINE_VERSION,
    arbitrate,
    assessment,
    canonical_story,
    compare_assessments,
    consistency_audit,
    criticize,
    evaluate_readiness,
    route_specialists,
    specialist_analysis,
)
from backend.model import GemmaRuntime
from backend.structured_output import generate_structured

#: How many of the 16 factors the model must score before its answer is accepted.
#: Below this the repair loop runs with the specific missing factors named.
MIN_MODEL_SCORED_FACTORS = 8

#: Context budget for the assembled story evidence, in characters. Kept modest because
#: prefill on a CPU-bound 1B model dominates wall-clock time.
STORY_CONTEXT_BUDGET = 6000


class Story(BaseModel):
    title: str = Field(min_length=1, max_length=500)
    user_story: str = Field(default="", max_length=20_000)
    acceptance_criteria: list[str] = Field(default_factory=list, max_length=50)
    technical_breakdown: str | None = Field(default=None, max_length=20_000)
    existing_points: float | None = None
    key: str | None = None
    status: str | None = None
    labels: list[str] = Field(default_factory=list)
    components: list[str] = Field(default_factory=list)
    source: Literal["manual", "jira", "upload"] = "manual"
    stack: StackProfile = Field(default_factory=StackProfile)

    @field_validator("acceptance_criteria", mode="before")
    @classmethod
    def normalize_criteria(cls, value: Any) -> list[str]:
        if value is None:
            return []
        if isinstance(value, str):
            return [item.strip() for item in value.replace(";", "\n").splitlines() if item.strip()]
        return [str(item).strip() for item in value if str(item).strip()]


class EstimateRequest(BaseModel):
    story: Story


class BatchEstimateRequest(BaseModel):
    stories: list[Story] = Field(min_length=1, max_length=100)


class JiraWriteRequest(BaseModel):
    points: Literal[1, 2, 3, 5, 8, 13, 21, 34]
    confirm: bool = False


class EffortRange(BaseModel):
    optimistic: float = Field(ge=0)
    likely: float = Field(ge=0)
    pessimistic: float = Field(ge=0)


class LayerEffort(BaseModel):
    frontend: str
    backend: str
    data: str
    assurance: str
    person_days: EffortRange


class HiddenTask(BaseModel):
    task: str
    weight: str


class Risk(BaseModel):
    risk: str
    mitigation_or_assumption: str


class SplitRecommendation(BaseModel):
    split_recommended: bool
    rationale: str
    proposed_stories: list[str] = Field(default_factory=list, max_length=6)


class EstimateDraft(BaseModel):
    """Tolerant boundary contract for a resource-constrained local model.

    Key casing, compact list forms, and partial scorecards are all accepted here and
    normalised below. A small model that returns nine good factor scores and skips seven
    is far more useful than one whose response is rejected wholesale.
    """

    scores: Any = Field(default_factory=dict)
    drivers: Any = Field(default_factory=list)
    points: Any = None
    rationale: Any = ""
    hidden_tasks: Any = Field(default_factory=list)
    risks: Any = Field(default_factory=list)
    assumptions: Any = Field(default_factory=list)
    proposed_stories: Any = Field(default_factory=list)

    @model_validator(mode="before")
    @classmethod
    def accept_compact_aliases(cls, value: Any) -> Any:
        if not isinstance(value, dict):
            return value
        normalized = {
            re.sub(r"[^a-z0-9]", "", str(key).lower()): item for key, item in value.items()
        }

        def pick(*names: str, default: Any = None) -> Any:
            for name in names:
                key = re.sub(r"[^a-z0-9]", "", name.lower())
                if key in normalized:
                    return normalized[key]
            return default

        split = pick("split_recommendation", "split", default={})
        split_data = split if isinstance(split, dict) else {}
        return {
            "scores": pick("scores", "scorecard", "factors", "factor_scores", default={}),
            "drivers": pick("drivers", "key_drivers", "complexity_drivers", default=[]),
            "points": pick("points", "story_points", "score"),
            "rationale": pick("rationale", "explanation", "reason", "why", default=""),
            "hidden_tasks": pick("hidden_tasks", "hiddentasks", "tasks", default=[]),
            "risks": pick("risks", "risk", default=[]),
            "assumptions": pick("assumptions", default=[]),
            "proposed_stories": split_data.get(
                "proposed_stories", pick("proposed_stories", "split_stories", default=[])
            ),
        }

    @model_validator(mode="after")
    def require_useful_signal(self) -> "EstimateDraft":
        if not self.scores and not self.drivers and self.points is None and not self.rationale:
            raise ValueError("At least one estimation signal is required")
        return self


# --------------------------------------------------------------------------------------
# Coercion helpers — every one of these exists because a 1B model shapes JSON loosely.
# --------------------------------------------------------------------------------------


def _items(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, (tuple, set)):
        return list(value)
    if isinstance(value, dict):
        return list(value.values())
    return [value]


def _text(value: Any, fallback: str = "") -> str:
    if value is None:
        return fallback
    if isinstance(value, dict):
        for key in ("reason", "why", "rationale", "note", "description", "text", "title", "name"):
            if value.get(key):
                return str(value[key]).strip()
        return fallback
    result = str(value).strip()
    return result or fallback


def _level(value: Any) -> int | None:
    """Coerce a model-supplied factor score to 1-5, or ``None`` if unusable."""
    if isinstance(value, dict):
        for key in ("score", "value", "level", "rating"):
            if key in value:
                value = value[key]
                break
        else:
            return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        words = {"none": 1, "trivial": 1, "low": 2, "minor": 2, "moderate": 3, "medium": 3,
                 "significant": 4, "high": 4, "extreme": 5, "very high": 5}
        cleaned = str(value).strip().lower()
        if cleaned in words:
            number = words[cleaned]
        else:
            match = re.search(r"\d+(?:\.\d+)?", cleaned)
            if not match:
                return None
            number = float(match.group())
    return max(1, min(5, round(number)))


def _canonical_factor(name: Any) -> str | None:
    """Map a loosely-written factor name onto a framework factor id."""
    key = re.sub(r"[^a-z0-9]+", "_", str(name).lower()).strip("_")
    if key in FACTOR_BY_ID:
        return key
    # Numeric keys ("1".."16") and label-ish keys ("technical complexity") both appear.
    if key.isdigit():
        number = int(key)
        for factor in FACTORS:
            if factor.number == number:
                return factor.id
        return None
    aliases = {
        "clarity": "requirements_clarity", "requirements": "requirements_clarity",
        "complexity": "technical_complexity", "technical": "technical_complexity",
        "integration": "integration_surface", "data_model": "data_model_change",
        "data": "data_model_change", "frontend": "frontend_effort", "ui": "frontend_effort",
        "backend": "backend_effort", "server": "backend_effort", "testing": "test_effort",
        "tests": "test_effort", "compliance": "regulatory_compliance",
        "regulatory": "regulatory_compliance", "security": "security_review",
        "observability": "observability_operations", "operations": "observability_operations",
        "cross_team": "cross_team_dependency", "dependencies": "cross_team_dependency",
        "rollback": "reversibility", "unknowns": "uncertainty", "risk": "uncertainty",
        "performance": "performance_scalability", "scalability": "performance_scalability",
        "documentation": "documentation_knowledge_transfer",
        "docs": "documentation_knowledge_transfer", "dod": "dod_overhead",
        "definition_of_done": "dod_overhead",
    }
    if key in aliases:
        return aliases[key]
    for factor in FACTORS:
        if key and (key in factor.id or factor.id in key):
            return factor.id
    return None


def _model_scores(draft: EstimateDraft) -> dict[str, dict[str, Any]]:
    """Extract ``factor id -> {score, reason}`` from whatever shape the model returned."""
    raw: dict[str, Any] = {}
    if isinstance(draft.scores, dict):
        raw = dict(draft.scores)
    else:
        for item in _items(draft.scores):
            if isinstance(item, dict):
                name = item.get("factor", item.get("parameter", item.get("name", "")))
                if name:
                    raw[str(name)] = item
    resolved: dict[str, dict[str, Any]] = {}
    for name, value in raw.items():
        factor_id = _canonical_factor(name)
        if not factor_id or factor_id in resolved:
            continue
        level = _level(value)
        if level is None:
            continue
        # A bare number ("uncertainty": 3) carries no reason. Stringifying it would put "3"
        # in the reason column, so only a mapping can supply explanatory text.
        reason = _text(value) if isinstance(value, dict) else ""
        resolved[factor_id] = {"score": level, "reason": reason}
    return resolved


# --------------------------------------------------------------------------------------
# Heuristic fallback — used only for factors the model did not score.
# --------------------------------------------------------------------------------------

_KEYWORDS: dict[str, tuple[str, ...]] = {
    "technical_complexity": ("algorithm", "concurren", "architect", "refactor", "async",
                             "distributed", "real-time", "realtime", "state machine"),
    "integration_surface": ("integrat", "third-party", "third party", "vendor", "external",
                            "webhook", "kafka", "queue", "api", "legacy"),
    "data_model_change": ("schema", "migration", "backfill", "database", "table", "index",
                          "entity", "persist"),
    "frontend_effort": ("ui", "screen", "form", "component", "responsive", "accessib",
                        "animation", "dashboard", "frontend"),
    "backend_effort": ("endpoint", "service", "backend", "business logic", "batch job",
                       "transaction", "worker", "cron"),
    "test_effort": ("test", "qa", "regression", "coverage", "e2e", "contract test",
                    "automation"),
    "regulatory_compliance": ("gdpr", "hipaa", "pci", "sox", "compliance", "regulat",
                              "audit trail", "retention", "consent"),
    "security_review": ("auth", "security", "encrypt", "pii", "token", "permission",
                        "biometric", "credential", "threat"),
    "observability_operations": ("monitor", "metric", "alert", "logging", "trace", "slo",
                                 "runbook", "dashboard", "on-call"),
    "cross_team_dependency": ("other team", "platform team", "external team", "depends on",
                              "blocked by", "provision", "shared"),
    "reversibility": ("migration", "deprecat", "irreversible", "cutover", "one-way",
                      "breaking change"),
    "performance_scalability": ("performance", "latency", "throughput", "scale", "load",
                                "cache", "concurrent users", "sla"),
    "documentation_knowledge_transfer": ("document", "adr", "runbook", "training",
                                         "knowledge", "handover", "public api"),
    "dod_overhead": ("release note", "demo", "rollout", "promotion", "sign-off",
                     "stakeholder", "launch", "deploy"),
}


def _story_evidence(story: Story) -> str:
    return " ".join(
        [
            story.title, story.user_story, " ".join(story.acceptance_criteria),
            story.technical_breakdown or "", " ".join(story.labels), " ".join(story.components),
        ]
    ).lower()


#: Verbs that describe an intention rather than a change. They are the strongest textual signal
#: of an unclear story, and they are independent of length — "investigate the vendor API" is six
#: words and wide open, while "change the button label to Log in" is seven words and closed.
_VAGUE = (
    "improve", "optimise", "optimize", "enhance", "review", "revisit", "refactor as needed",
    "support", "handle", "as required", "etc", "and so on", "tbd", "tbc", "somehow",
    "if needed", "where possible",
)

#: Stories that ask for an answer rather than a change. These are different in kind from merely
#: vague ones: "improve performance" is an under-specified change, but "investigate whether the
#: vendor API works" has no known implementation path at all, which is the definition of maximum
#: uncertainty and the case the spike gate exists for.
_EXPLORATORY = (
    "investigate", "explore", "research", "look into", "figure out", "find out whether",
    "assess whether", "spike", "proof of concept", "feasibility", "evaluate whether",
)


#: Marks of a story that closes its own scope: a quoted literal, a from/to, an identifier, a
#: number, a file or field name. Short is not the same as unspecified — "rename `customerName`
#: to `customer_name`" is seven words and completely bounded.
_CONCRETE = re.compile(
    r"""['"`][^'"`]{2,}['"`]"""      # a quoted literal
    r"""|\bfrom\s+\S+\s+to\s+\S+"""  # a stated transition
    r"""|\b\w+[._]\w+\b"""           # an identifier: customer_name, orders.status
    r"""|\b\d+\b"""                  # a number: 20 per page, 500ms
    r"""|\b[a-z]+[A-Z]\w*\b"""       # camelCase: customerName
)


def _story_specified(story: Story, evidence: str) -> bool:
    """Does the story pin down its own finished state?

    This, not length, is what licenses a low score. A story that states what "done" looks like
    can be scored small on the factors it does not mention, because it has bounded itself. A
    story that does not — "improve reporting", "support the new vendor" — has said nothing
    about scope, and reading that silence as simplicity is how an unbounded piece of work gets
    a confident small number.
    """
    if story.acceptance_criteria or story.technical_breakdown:
        return True
    # A concrete marker still needs a sentence around it: a bare number in three words is
    # not a specification.
    return bool(_CONCRETE.search(evidence)) and len(evidence.strip()) >= 24


def _story_scale(story: Story, evidence: str) -> int:
    """0-3: how much work the *shape* of the story implies, before any factor is considered.

    Needed because an absence of evidence means opposite things at opposite sizes. A story of
    twelve words that never mentions testing probably has very little; a story with eight
    acceptance criteria across three components that never mentions testing has plenty — the
    story just failed to say so. Scoring both at the same baseline is what made every estimate
    land in the same band.
    """
    length = len(evidence.strip())
    signals = (
        length >= 220,
        length >= 700,
        len(story.acceptance_criteria) >= 3,
        len(story.acceptance_criteria) >= 6,
        bool(story.technical_breakdown),
        len(story.components) >= 2,
    )
    return min(3, sum(signals))


def _heuristic_score(factor_id: str, story: Story, evidence: str) -> tuple[int, str]:
    """Derive a defensible 1-5 score from story text when the model skipped a factor."""
    stack = story.stack
    scale = _story_scale(story, evidence)
    vague = [term for term in _VAGUE if term in evidence]
    exploratory = [term for term in _EXPLORATORY if term in evidence]

    if factor_id == "requirements_clarity":
        count = len(story.acceptance_criteria)
        if exploratory:
            return 5, (
                f"The story asks to {exploratory[0]} — there is no described finished state to "
                f"build against."
            )
        if vague:
            return (
                min(5, 3 + (scale == 0)),
                f"The story asks to {vague[0]} rather than describing a finished state.",
            )
        if count >= 3 and len(evidence) > 240:
            return 2, f"{count} acceptance criteria and a substantive description are present."
        if count >= 1:
            return 3, f"Only {count} acceptance criterion/criteria supplied; gaps are likely."
        # A short, self-contained change with no criteria is not ambiguous — it is small. The
        # previous rule scored it 4, which is why a one-line copy change came out the same size
        # as a migration.
        if scale == 0:
            return 2, "No criteria, but the change is small and states its own finished state."
        return 4, "No acceptance criteria were supplied, so the requirement is not pinned down."

    if factor_id == "uncertainty":
        if stack.maturity_level == 5:
            return 5, "The declared framework is bleeding-edge; unknowns dominate."
        if exploratory:
            # Maximum uncertainty by definition: the story is asking what the work is. The
            # framework's spike gate keys on 5, so this is what stops the pipeline handing back
            # a confident number for a question nobody has answered yet.
            return 5, (
                f"The story asks to {exploratory[0]} rather than to build something; the "
                f"implementation path is unknown by design."
            )
        if vague:
            return 4, f"The story is exploratory ('{vague[0]}'), so the path is not yet known."
        if stack.scenario in {"new_framework", "framework_upgrade"}:
            return 4, f"The declared scenario ({stack.scenario.replace('_', ' ')}) carries unknowns."
        if not story.acceptance_criteria and scale == 0:
            return 3, "Sparse story evidence, though the change itself is small."
        if scale >= 2 and story.technical_breakdown:
            return 2, "The story is substantial but the implementation path is described."
        if scale == 0:
            return 2, "A small, self-contained change with little room for surprise."
        return 3, "Evidence exists, but implementation unknowns have not been ruled out."

    if factor_id == "frontend_effort" and stack.frontend == "none":
        return 1, "No frontend stack is declared for this story."
    if factor_id == "backend_effort" and stack.backend == "none":
        return 1, "No backend stack is declared for this story."

    terms = _KEYWORDS.get(factor_id, ())
    matched = [term for term in terms if term in evidence]
    label = FACTOR_BY_ID[factor_id].label.lower()
    if not matched:
        # Absence of evidence is not evidence of absence. A story that has bounded itself may be
        # scored low on what it does not mention; a story that has not says nothing about scope,
        # and scoring that silence as "small" is the failure this whole framework exists to
        # avoid. Either way the reason states which of the two happened, because a 1 that means
        # "the story rules this out" and a 4 that means "the story never said" are different
        # claims and a reader has to be able to tell them apart.
        if not _story_specified(story, evidence):
            return 4, (
                f"The story does not say whether {label} is involved, and it does not state "
                f"what done looks like. Scored high because unstated scope is unbounded, not "
                f"because evidence was found."
            )
        if scale == 0:
            return 1, (
                f"The story states its finished state and no {label} work follows from it."
            )
        if scale >= 3:
            return 3, (
                f"The story is large and says nothing about {label}; unstated work at this size "
                f"is more likely to exist than not."
            )
        return 2, (
            f"The story bounds its scope and gives no {label} evidence; scored at baseline for "
            f"a story this size."
        )

    # Matches drive the score; the size of the story nudges it by at most one, so a single
    # incidental keyword in a long story cannot reach the top of the scale on its own.
    score = min(5, 1 + min(3, len(matched)) + (1 if scale >= 2 else 0))
    sample = ", ".join(sorted(matched)[:3])
    return score, f"Story evidence mentions {sample}."


# --------------------------------------------------------------------------------------
# Scorecard assembly
# --------------------------------------------------------------------------------------


def build_scorecard(draft: EstimateDraft, story: Story) -> list[FactorScore]:
    """Merge model judgement with heuristic fallback into all 16 factors, in order."""
    supplied = _model_scores(draft)
    evidence = _story_evidence(story)
    guidance = story.stack.guidance()
    scorecard: list[FactorScore] = []
    for factor in FACTORS:
        provided = supplied.get(factor.id)
        provenance: Literal["model", "heuristic"] = "model" if provided else "heuristic"
        if provided:
            score = int(provided["score"])
            # Small models often answer with a bare number. Rather than showing an empty
            # cell, fall back to the keyword evidence for the same factor — it describes
            # the same story text, and the score stays attributed to the model.
            reason = provided["reason"] or _heuristic_score(factor.id, story, evidence)[1]
        else:
            score, reason = _heuristic_score(factor.id, story, evidence)
        scorecard.append(
            FactorScore(
                factor=factor.id, number=factor.number, label=factor.label,
                group=factor.group, score=score, reason=reason[:300], provenance=provenance,
                stack_notes=guidance.get(factor.id, []),
            )
        )
    return scorecard


def _effort(points: int, scores: dict[str, int]) -> LayerEffort:
    optimistic, likely, pessimistic = EFFORT_DEFAULTS[points]
    words = {1: "None", 2: "Small", 3: "Moderate", 4: "Significant", 5: "Extensive"}
    return LayerEffort(
        frontend=f"{words[scores['frontend_effort']]} frontend scope",
        backend=f"{words[scores['backend_effort']]} backend scope",
        data=f"{words[scores['data_model_change']]} data-model scope",
        assurance=(
            f"{words[max(scores['test_effort'], scores['security_review'])]} "
            "test and security scope"
        ),
        person_days=EffortRange(optimistic=optimistic, likely=likely, pessimistic=pessimistic),
    )


def _drivers(draft: EstimateDraft, scorecard: list[FactorScore]) -> list[str]:
    """The 2-3 factors actually moving the number, ranked by score then framework order."""
    ranked = sorted(scorecard, key=lambda item: (-item.score, item.number))
    drivers = [f"{item.label} ({item.score})" for item in ranked[:3] if item.score >= 3]
    if len(drivers) < 2:
        drivers = [f"{item.label} ({item.score})" for item in ranked[:2]]
    return drivers


def _collect(values: Any, limit: int) -> list[str]:
    return [text for text in (_text(item) for item in _items(values)) if text][:limit]


def build_result(
    draft: EstimateDraft,
    story: Story,
    context_manifest: list[dict],
    scorecard_override: list[FactorScore] | None = None,
) -> dict[str, Any]:
    """Turn a model draft plus the framework arithmetic into the full estimate payload."""
    scorecard = scorecard_override or build_scorecard(draft, story)
    scores = {item.factor: int(item.score) for item in scorecard}
    stack = story.stack

    calculation: Calculation = calculate(scores, stack)
    checks = policy_checks(scores, stack, calculation)
    recommendation, recommendation_detail = decide(checks, stack, calculation, scores)
    confidence_level, confidence_detail = confidence(scores, stack, calculation)
    reasoning = detailed_reasoning(
        scorecard,
        stack,
        calculation,
        checks,
        recommendation,
        recommendation_detail,
        confidence_detail,
    )
    suggestions = estimation_suggestions(
        scorecard, stack, calculation, checks, recommendation, reasoning
    )
    flags = risk_flags(scores, stack)
    anchors = stack.anchors()
    nearest = min(anchors, key=lambda anchor: abs(int(anchor["points"]) - calculation.points))

    # The model's own guess is kept purely as a cross-check signal. Divergence is shown to
    # the reader rather than resolved silently in either direction.
    model_points_guess = None
    if draft.points is not None:
        match = re.search(r"\d+", str(draft.points))
        if match:
            candidate = int(match.group())
            model_points_guess = min(FIBONACCI_POINTS, key=lambda p: abs(p - candidate))

    spike_needed = recommendation in {"spike_first", "upgrade_framework_first", "epic_discovery"}
    split_needed = recommendation in {"decompose", "epic_discovery"}
    unknowns = [item.label for item in scorecard if item.score >= 4]
    proposed = _collect(draft.proposed_stories, 6)

    hidden_tasks: list[HiddenTask] = []
    for item in _items(draft.hidden_tasks)[:8]:
        data = item if isinstance(item, dict) else {}
        task = _text(data.get("task", data.get("title", item)))
        if task:
            hidden_tasks.append(
                HiddenTask(task=task, weight=_text(data.get("weight"), "Supporting work"))
            )
    for item in scorecard:
        if len(hidden_tasks) >= 8:
            break
        # Assurance work at 4+ is the classic source of "we forgot that" overruns, so it
        # is surfaced as explicit hidden work even when the model did not mention it.
        if item.score >= 4 and item.factor in {
            "test_effort", "documentation_knowledge_transfer", "observability_operations",
            "dod_overhead", "regulatory_compliance",
        } and not any(item.label.lower() in task.task.lower() for task in hidden_tasks):
            hidden_tasks.append(
                HiddenTask(task=item.label, weight=f"Scored {item.score}/5 — {item.reason}")
            )

    risks: list[Risk] = []
    for item in _items(draft.risks)[:3]:
        data = item if isinstance(item, dict) else {}
        risk = _text(data.get("risk", data.get("title", item)))
        if risk:
            risks.append(
                Risk(
                    risk=risk,
                    mitigation_or_assumption=_text(
                        data.get("mitigation_or_assumption", data.get("mitigation")),
                        "Validate this risk before the delivery commitment.",
                    ),
                )
            )
    for flag in flags:
        if len(risks) >= 4:
            break
        if not any(str(flag["label"]).lower() in item.risk.lower() for item in risks):
            risks.append(
                Risk(risk=str(flag["label"]), mitigation_or_assumption=str(flag["detail"]))
            )
    if not risks:
        risks = [
            Risk(
                risk="Estimate uncertainty",
                mitigation_or_assumption="Confirm assumptions and acceptance criteria first.",
            )
        ]

    assumptions = _collect(draft.assumptions, 8) or [
        "The estimate uses only the supplied story evidence and the declared stack profile."
    ]
    rationale = _text(draft.rationale) or (
        f"{calculation.base_sum} base points across 16 factors, adjusted by "
        f"{calculation.base_adjustment_total + calculation.stack_adjustment_total}, "
        f"lands in band {calculation.band}."
    )
    drivers = _drivers(draft, scorecard)
    tldr = (
        f"{calculation.points} points — {', '.join(drivers[:2])} drive the estimate. "
        f"{confidence_level} confidence."
    )

    maturity = MATURITY_TAXONOMY[int(stack.maturity_level)]
    model_scored = sum(1 for item in scorecard if item.provenance == "model")

    return {
        "framework": {
            "name": "Agile Story Point Estimation Framework",
            "version": FRAMEWORK_VERSION,
            "document": FRAMEWORK_DOCUMENT,
            "factor_count": len(FACTORS),
        },
        "story": story.model_dump(),
        "stack": {
            **stack.model_dump(),
            "frontend_label": STACK_LABELS.get(stack.frontend, stack.frontend),
            "backend_label": STACK_LABELS.get(stack.backend, stack.backend),
            "maturity_name": maturity["name"],
            "maturity_definition": maturity["definition"],
            "maturity_action": maturity["action"],
        },
        "scorecard": [item.model_dump() for item in scorecard],
        "calculation": calculation.model_dump(),
        "points": calculation.points,
        "drivers": drivers,
        "drivers_explanation": rationale,
        "tldr": tldr,
        "plain_language_why": (
            f"{recommendation_detail} {confidence_detail} "
            f"The score is driven by {', '.join(drivers[:2])}."
        ),
        "confidence": confidence_level,
        "confidence_detail": confidence_detail,
        "recommendation": recommendation,
        "recommendation_detail": recommendation_detail,
        "detailed_reasoning": reasoning.model_dump(),
        "suggestions": [item.model_dump() for item in suggestions],
        "risk_flags": flags,
        "anchor_comparison": (
            f"Closest calibrated reference: \"{nearest['title']}\" "
            f"({nearest['stack']}, {nearest['points']} points)."
        ),
        "anchors_considered": anchors,
        "effort": _effort(calculation.points, scores).model_dump(),
        "hidden_tasks": [item.model_dump() for item in hidden_tasks],
        "risks": [item.model_dump() for item in risks],
        "assumptions": assumptions,
        "spike_recommended": spike_needed,
        "spike_reason": recommendation_detail if spike_needed else None,
        "spike_definition": spike_template(story.title, unknowns) if spike_needed else None,
        "split_recommendation": SplitRecommendation(
            split_recommended=split_needed,
            rationale=(
                recommendation_detail if split_needed
                else "The story is cohesive at its calculated size."
            ),
            proposed_stories=proposed,
        ).model_dump(),
        "evidence": {
            "source": story.source,
            "context_manifest": context_manifest,
            "policy_checks": [check.model_dump() for check in checks],
            "scoring_provenance": {
                "model_scored": model_scored,
                "heuristic_filled": len(scorecard) - model_scored,
                "minimum_required": MIN_MODEL_SCORED_FACTORS,
            },
            "model_cross_check": {
                "model_points": model_points_guess,
                "calculated_points": calculation.points,
                "agreement": (
                    "not_offered" if model_points_guess is None
                    else "agrees" if model_points_guess == calculation.points
                    else "diverges"
                ),
                "note": (
                    "The reported number comes from the framework arithmetic. The model's own "
                    "guess is shown only so disagreement is visible."
                ),
            },
            "determinism": (
                "Story points are computed from the scorecard by fixed rules. Re-running the "
                "arithmetic on the same scores always yields the same points."
            ),
        },
    }


# --------------------------------------------------------------------------------------
# Prompt construction and the estimation service
# --------------------------------------------------------------------------------------

_SYSTEM = (
    "You are a senior technical estimator running a calibrated multi-factor analysis. "
    "You score evidence; you never invent requirements, and you never compute the story "
    "points yourself. If the story does not evidence a factor, score it low and say so."
)


def _rubric(stack: StackProfile) -> str:
    """Render the 16-factor rubric, injecting only the declared stack's guidance."""
    guidance = stack.guidance()
    lines = []
    for factor in FACTORS:
        line = f"{factor.number}. {factor.id} — {factor.description} 1 = {factor.low_anchor} 5 = {factor.high_anchor}"
        notes = guidance.get(factor.id)
        if notes:
            line += "\n   Stack calibration: " + " | ".join(notes)
        lines.append(line)
    return "\n".join(lines)


def build_prompt(story: Story) -> tuple[str, list[dict]]:
    """Assemble the estimation prompt and return it with its provenance manifest.

    Story text is marked untrusted in the context envelope: it arrives from Jira, a
    spreadsheet, or a paste box, and must be read as evidence rather than as instructions.
    """
    stack = story.stack
    anchors = "\n".join(
        f"- {anchor['points']} points ({anchor['stack']}): {anchor['title']}"
        for anchor in stack.anchors()
    )
    sources = [
        ContextSource(
            id="story", label="Story under estimation", priority=100, trusted=False,
            content=json.dumps(
                {
                    "title": story.title,
                    "user_story": story.user_story,
                    "acceptance_criteria": story.acceptance_criteria,
                    "technical_breakdown": story.technical_breakdown,
                    "labels": story.labels,
                    "components": story.components,
                },
                indent=2,
            ),
        ),
        ContextSource(
            id="stack_profile", label="Declared technology stack", priority=90, trusted=True,
            content=json.dumps(
                {
                    "frontend": STACK_LABELS.get(stack.frontend, stack.frontend),
                    "backend": STACK_LABELS.get(stack.backend, stack.backend),
                    "database": stack.database or "unspecified",
                    "framework_maturity": f"{stack.maturity_level} "
                    f"({MATURITY_TAXONOMY[int(stack.maturity_level)]['name']})",
                    "team_experience": stack.team_experience,
                    "scenario": stack.scenario,
                },
                indent=2,
            ),
        ),
        ContextSource(
            id="anchors", label="Calibrated reference stories", priority=80, trusted=True,
            content=anchors,
        ),
    ]
    context, manifest = assemble_context(sources, STORY_CONTEXT_BUDGET)
    prompt = f"""Score the story below against all 16 factors of the estimation framework.

{context}

FACTOR RUBRIC — score every factor from 1 to 5:
{_rubric(stack)}

{GROUNDING_CONTRACT_BRIEF}

Rules:
- Score all 16 factors using their exact ids.
- Score only what the story says. Do not invent requirements, components, integrations, or
  constraints it does not state, and do not assume a technology it does not name.
- Silence is not evidence that the work is small. If the story does not give you enough to
  judge a factor, score it 4 and say so in the reason — for example "the story does not say
  whether existing data must be migrated". Unstated scope is unbounded scope, and an estimate
  that reads absence as simplicity is how a two-line story becomes a two-week surprise.
- Score a factor 1 or 2 only when the story positively bounds it: it states the finished state,
  names the files or screens involved, or the change is self-evidently closed.
- Give each score a reason of at most 25 words, quoting or paraphrasing the story evidence, or
  naming exactly what the story failed to say.
- Name the 2-3 factors that genuinely drive the size.
- List hidden sub-tasks the story text omits but the work implies.

Return one JSON object with these keys:
  scores: object mapping each factor id to {{"score": 1-5, "why": "short reason"}}
  drivers: array of 2-3 factor ids
  rationale: one sentence explaining the overall size
  hidden_tasks: array of {{"task": ..., "weight": ...}}
  risks: array of {{"risk": ..., "mitigation_or_assumption": ...}}
  assumptions: array of strings
  proposed_stories: array of smaller story titles, only if this should be split
"""
    return prompt, manifest


def _validate_draft(draft: EstimateDraft) -> str | None:
    """Semantic gate for the repair loop: name exactly what is missing, not just 'invalid'.

    Feeding back the specific unscored factor ids is what makes the second attempt useful
    on a small model; a generic "try again" tends to reproduce the same omissions.
    """
    supplied = _model_scores(draft)
    if len(supplied) >= MIN_MODEL_SCORED_FACTORS:
        return None
    missing = [factor_id for factor_id in FACTOR_IDS if factor_id not in supplied]
    return (
        f"Only {len(supplied)} of 16 factors were scored. Score at least "
        f"{MIN_MODEL_SCORED_FACTORS}. These are still missing: {', '.join(missing)}. "
        'Use the exact ids as keys, each mapping to {"score": 1-5, "why": "..."}.'
    )


#: The blind reviewer runs warmer than the primary pass. At a shared low temperature both
#: passes converge on nearly identical scores, so the second generation costs minutes of CPU
#: and produces no independent signal to arbitrate between.
BLIND_REVIEW_TEMPERATURE = 0.7

#: How close to a Fibonacci band edge counts as "a different opinion could move this".
BAND_EDGE_MARGIN = 3


def blind_review_warranted(primary, story: Story) -> tuple[bool, str]:
    """Decide whether a second independent pass can change the outcome.

    A second full generation roughly doubles the wall-clock cost of an estimate on a CPU
    model. It is worth that when the answer is genuinely in play — near a band edge, on a
    protected risk dimension, or where the primary pass was mostly guessing — and is not
    worth it for a small, well-evidenced story sitting in the middle of its band.

    Returns the decision and the reason, so the UI can show why it was or was not run.
    """
    scores = {item.factor: int(item.score_most_likely) for item in primary.dimensions}
    calculation = calculate(scores, story.stack)

    lower, _, upper = _band_edges(calculation.adjusted_score)
    distance = min(calculation.adjusted_score - lower, upper - calculation.adjusted_score)
    if distance <= BAND_EDGE_MARGIN:
        return True, (
            f"Adjusted score {calculation.adjusted_score} sits {distance} from a band edge, "
            "so a different opinion could change the points."
        )

    elevated = [FACTOR_BY_ID[key].label for key, value in scores.items() if value >= 4]
    protected = [
        FACTOR_BY_ID[key].label
        for key, value in scores.items()
        if value >= 4 and key in PROTECTED_FACTORS
    ]
    if protected:
        return True, f"Protected risk dimension elevated: {', '.join(protected)}."
    if len(elevated) >= 3:
        return True, f"{len(elevated)} factors scored 4 or above."
    if primary.heuristic_filled > primary.model_scored:
        return True, (
            f"{primary.heuristic_filled} of 16 factors were inferred rather than scored, "
            "so a second reading is worth the time."
        )
    if int(story.stack.maturity_level) >= 4 or int(story.stack.team_experience) <= 2:
        return True, "The declared stack carries a maturity or experience penalty."

    return False, (
        f"Adjusted score {calculation.adjusted_score} is {distance} from either band edge with "
        "no elevated risk factor, so a second pass cannot change the result."
    )


def _band_edges(score: int) -> tuple[int, int, int]:
    """Inclusive lower bound, the score, and inclusive upper bound of its Fibonacci band."""
    bounds = [(16, 24), (25, 34), (35, 44), (45, 54), (55, 64), (65, 80)]
    for lower, upper in bounds:
        if score <= upper:
            return max(lower, 16), score, upper
    return 65, score, 80


class EstimateService:
    def __init__(self, runtime: GemmaRuntime, settings: Settings):
        self.runtime = runtime
        self.settings = settings

    async def estimate(
        self,
        story: Story,
        progress: Callable[[dict[str, Any]], None] | None = None,
        reference_history: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        # EAGLE §2: the contract is fixed before anything reads the story, and every later
        # stage works against it rather than against a story someone can still edit.
        contract = build_contract(story)
        if progress:
            progress(
                {
                    "stage": "contract",
                    "status": "completed",
                    "label": f"Estimation contract {contract.contract_hash[7:15]} sealed",
                    "detail": (
                        "Objective, acceptance criteria, stack, completion rules and stop "
                        "conditions are frozen for the run."
                    ),
                    "evidence": {
                        "story_id": contract.story_id,
                        "required_evidence": list(contract.required_evidence()),
                        "stop_conditions": contract.stop_conditions.model_dump(),
                        "max_debate_rounds": contract.max_debate_rounds,
                    },
                }
            )
        canonical = canonical_story(story)
        readiness = evaluate_readiness(story, canonical)
        pipeline_mode, specialist_routes = route_specialists(story)
        if progress:
            progress(
                {
                    "stage": "normalize",
                    "status": "completed",
                    "label": "Story normalized into stable evidence",
                    "evidence": {
                        "input_hash": canonical["input_hash"],
                        "evidence_items": len(canonical["evidence"]),
                        "missing_fields": canonical["missing_fields"] or "none",
                        "untrusted_instructions_detected": canonical[
                            "untrusted_instructions_detected"
                        ],
                    },
                }
            )
            progress(
                {
                    "stage": "readiness",
                    "status": "completed",
                    "label": f"Readiness: {readiness.decision.replace('_', ' ').title()}",
                    "evidence": {
                        "checks": len(readiness.checks),
                        "assumptions": len(readiness.assumptions),
                        "questions": len(readiness.targeted_questions),
                    },
                }
            )
            progress(
                {
                    "stage": "specialist_routing",
                    "status": "completed",
                    "label": (
                        f"{pipeline_mode.replace('_', ' ').title()} pipeline routed to "
                        f"{len(specialist_routes)} specialist lenses"
                    ),
                    "evidence": {
                        "mode": pipeline_mode,
                        "specialists": [route.label for route in specialist_routes],
                    },
                }
            )
        prompt, manifest = build_prompt(story)
        if progress:
            progress(
                {
                    "stage": "assemble_context",
                    "status": "completed",
                    "label": "Story evidence bounded and labelled",
                    "evidence": {
                        "sources": len(manifest),
                        "characters": sum(int(item["characters"]) for item in manifest),
                        "budget": STORY_CONTEXT_BUDGET,
                        "truncated": any(item["truncated"] for item in manifest),
                        "untrusted_sources": sum(1 for item in manifest if not item["trusted"]),
                    },
                }
            )
            progress(
                {
                    "stage": "declare_stack",
                    "status": "completed",
                    "label": (
                        f"Stack calibration loaded for "
                        f"{STACK_LABELS.get(story.stack.frontend, 'None')} / "
                        f"{STACK_LABELS.get(story.stack.backend, 'None')}"
                    ),
                    "evidence": {
                        "maturity": (
                            f"{story.stack.maturity_level} "
                            f"({MATURITY_TAXONOMY[int(story.stack.maturity_level)]['name']})"
                        ),
                        "team_experience": story.stack.team_experience,
                        "scenario": story.stack.scenario,
                        "reference_anchors": len(story.stack.anchors()),
                    },
                }
            )

        try:
            draft = await generate_structured(
                self.runtime,
                EstimateDraft,
                _SYSTEM,
                prompt,
                max_new_tokens=self.settings.estimate_max_output_tokens,
                validate_result=_validate_draft,
                on_attempt=(
                    lambda event: progress({"stage": "primary_estimate", **event})
                    if progress else None
                ),
            )
        except ValueError as exc:
            # A small local model that cannot hold the contract must not cost the user their
            # estimate. The heuristic scorecard still produces a defensible number, and the
            # degradation is reported rather than hidden.
            if progress:
                progress(
                    {
                        "stage": "primary_estimate",
                        "status": "failed",
                        "label": "Model output unusable — falling back to evidence heuristics",
                        "detail": str(exc)[:500],
                    }
                )
            draft = EstimateDraft.model_construct(
                scores={}, drivers=[], points=None, rationale="",
                hidden_tasks=[], risks=[], assumptions=[], proposed_stories=[],
            )
        primary_scorecard = build_scorecard(draft, story)
        primary_assessment = assessment("PRIMARY_ESTIMATOR", primary_scorecard, story)
        specialist_findings = specialist_analysis(specialist_routes, primary_assessment)
        if progress:
            progress(
                {
                    "stage": "primary_estimate",
                    "status": "completed",
                    "label": "Primary evidence assessment completed",
                    "evidence": {
                        "model_scored": primary_assessment.model_scored,
                        "heuristic_filled": primary_assessment.heuristic_filled,
                        "point_cross_check": primary_assessment.point_cross_check,
                    },
                }
            )
            progress(
                {
                    "stage": "specialist_analysis",
                    "status": "completed",
                    "label": f"{len(specialist_findings)} specialist lenses assessed evidence",
                    "evidence": {
                        "lenses": [item.label for item in specialist_findings],
                        "material_risks": sum(
                            len(item.material_risks) for item in specialist_findings
                        ),
                        "open_questions": sum(
                            len(item.open_questions) for item in specialist_findings
                        ),
                    },
                }
            )

        # A second full generation doubles the wall-clock cost of an estimate on CPU, so it
        # runs where a second opinion can actually change the answer rather than on every
        # story. `review_reason` records which test triggered it, so the decision is visible.
        review_needed, review_reason = blind_review_warranted(primary_assessment, story)
        if not review_needed and progress:
            progress(
                {
                    "stage": "blind_review",
                    "status": "completed",
                    "label": "Blind review not required for this story",
                    "detail": review_reason,
                    "evidence": {"executed": False, "reason": review_reason},
                }
            )

        reviewer_draft: EstimateDraft | None = None
        if review_needed:
            reviewer_system = (
                "You are an independent blind technical estimator. You have not seen another "
                "estimator's scores. Assess only the supplied evidence against the rubric. "
                "Score every factor from 1 to 5, do not invent requirements, and never compute "
                "or recommend story points. Keep each reason concise and evidence-specific."
            )
            try:
                reviewer_draft = await generate_structured(
                    self.runtime,
                    EstimateDraft,
                    reviewer_system,
                    prompt,
                    max_new_tokens=self.settings.estimate_max_output_tokens,
                    validate_result=_validate_draft,
                    # Deliberately warmer than the primary pass. At the shared default
                    # temperature both passes converge on nearly the same scores, which
                    # costs a full generation and yields no independent signal.
                    temperature=max(self.settings.temperature, BLIND_REVIEW_TEMPERATURE),
                    on_attempt=(
                        lambda event: progress({"stage": "blind_review", **event})
                        if progress else None
                    ),
                )
            except ValueError as exc:
                if progress:
                    progress(
                        {
                            "stage": "blind_review",
                            "status": "failed",
                            "label": "Blind review degraded to independent evidence heuristics",
                            "detail": str(exc)[:500],
                        }
                    )
        blind_review_executed = reviewer_draft is not None
        if blind_review_executed:
            reviewer_scorecard = build_scorecard(reviewer_draft, story)
        else:
            # Skipped or degraded. Mirroring the primary is deliberate: the alternative is to
            # score the reviewer from keyword heuristics, which is not an independent opinion
            # but a fallback for *missing* scores. Arbitrating against it manufactures
            # disagreement — on one story it read "no frontend declared" as 1 against the
            # model's 3 and quietly moved the final score — so not running the review would
            # change the estimate, which is worse than either running it or skipping it.
            reviewer_scorecard = primary_scorecard
        reviewer_assessment = assessment("BLIND_REVIEWER", reviewer_scorecard, story)
        if progress:
            progress(
                {
                    "stage": "blind_review",
                    "status": "completed",
                    "label": "Independent blind review completed",
                    "detail": "The reviewer received story evidence and rubric, never primary scores.",
                    "evidence": {
                        "blind": True,
                        "executed": blind_review_executed,
                        "reason": review_reason,
                        "model_scored": reviewer_assessment.model_scored,
                        "heuristic_filled": reviewer_assessment.heuristic_filled,
                        "point_cross_check": reviewer_assessment.point_cross_check,
                    },
                }
            )

        disagreements = compare_assessments(primary_assessment, reviewer_assessment)
        challenges = criticize(disagreements)
        arbitrated_scores, arbitration = arbitrate(
            primary_assessment, reviewer_assessment, disagreements
        )

        # -- EAGLE governance ---------------------------------------------------------------
        # The blackboard is built from what the scorers actually cited, so a heuristic fill is
        # recorded as a low-confidence claim rather than passed off as something read from the
        # story. Everything downstream cites these records by id.
        board = build_blackboard(story, primary_scorecard)
        proposals = [
            {item.factor: item.score_most_likely for item in primary_assessment.dimensions},
            {item.factor: item.score_most_likely for item in reviewer_assessment.dimensions},
        ]
        medians, aggregates = aggregate(proposals, board)
        eagle_findings = (
            critic_review(aggregates, medians, story.stack, board)
            + adversarial_review(medians, story.stack, board, _story_evidence(story))
            + optimistic_review(medians, story.stack, board)
        )
        resolved_scores, debate_outcome = debate(aggregates, eagle_findings, contract)
        # The resolved scores carry the same shape the framework expects, with the reason kept
        # from whichever proposal the debate settled on.
        arbitrated_scores = {
            factor: {
                "score": score,
                "why": next(
                    (row.reason for row in aggregates if row.factor == factor), "median accepted"
                ),
            }
            for factor, score in resolved_scores.items()
        }
        if progress:
            disputed = [row for row in aggregates if row.status == "dispute"]
            progress(
                {
                    "stage": "eagle_conflict",
                    "status": "completed" if not disputed else "retrying",
                    "label": (
                        f"{len(disputed)} factor(s) disputed on spread or missing evidence"
                        if disputed else "Independent proposals agree on every factor"
                    ),
                    "detail": (
                        "Spread of 0 accepts, 1 accepts the median, 2 or more disputes; an "
                        "elevated score with no evidence disputes regardless of agreement."
                    ),
                    "evidence": {
                        "disputed": [row.label for row in disputed],
                        "owners": sorted({row.owner for row in disputed}),
                        "estimator_count": len(proposals),
                    },
                }
            )
            progress(
                {
                    "stage": "eagle_review",
                    "status": "completed",
                    "label": (
                        f"Critic, adversarial and optimistic reviewers raised "
                        f"{len(eagle_findings)} finding(s)"
                    ),
                    "detail": (
                        "The adversarial reviewer looks only for under-estimation; the "
                        "optimistic reviewer only for complexity counted twice."
                    ),
                    "evidence": {
                        "blocker": sum(i.severity == "blocker" for i in eagle_findings),
                        "material": sum(i.severity == "material" for i in eagle_findings),
                        "advisory": sum(i.severity == "advisory" for i in eagle_findings),
                        "by_reviewer": {
                            name: sum(i.reviewer == name for i in eagle_findings)
                            for name in ("critic", "adversarial", "optimistic")
                        },
                    },
                }
            )
            if debate_outcome.rounds:
                progress(
                    {
                        "stage": "eagle_debate",
                        "status": (
                            "waiting" if debate_outcome.escalation == "HUMAN_REVIEW"
                            else "completed"
                        ),
                        "label": (
                            f"Targeted debate over {len(debate_outcome.factors_debated)} "
                            f"disputed factor(s) in {len(debate_outcome.rounds)} round(s)"
                        ),
                        "detail": (
                            "Only disputed factors are re-examined; the rest of the pipeline "
                            "is not re-run."
                        ),
                        "evidence": {
                            "factors": debate_outcome.factors_debated,
                            "unresolved": debate_outcome.unresolved,
                            "escalation": debate_outcome.escalation,
                            "max_rounds": contract.max_debate_rounds,
                        },
                    }
                )
        if progress:
            material = sum(item.material for item in disagreements)
            progress(
                {
                    "stage": "disagreement",
                    "status": "completed",
                    "label": f"Independent estimates compared: {material} material disagreement(s)",
                    "evidence": {
                        "differences": len(disagreements),
                        "material": material,
                        "protected": sum(
                            item.material and item.protected for item in disagreements
                        ),
                    },
                }
            )
            progress(
                {
                    "stage": "critic",
                    "status": "completed",
                    "label": (
                        f"Critic challenged {len(challenges)} material dimension(s)"
                        if challenges else "Critic found no material challenge"
                    ),
                    "evidence": {"challenges": len(challenges)},
                }
            )
            progress(
                {
                    "stage": "arbitration",
                    "status": "completed",
                    "label": "Disagreements resolved by explicit deterministic policy",
                    "evidence": {
                        "decisions": len(arbitration),
                        "human_approval_required": sum(
                            item.human_approval_required for item in arbitration
                        ),
                    },
                }
            )

        final_draft = EstimateDraft.model_construct(
            scores=arbitrated_scores,
            drivers=draft.drivers,
            points=draft.points,
            rationale=(
                "Primary and blind assessments were compared and reconciled by the "
                "published arbitration policy."
            ),
            hidden_tasks=draft.hidden_tasks,
            risks=draft.risks,
            assumptions=draft.assumptions,
            proposed_stories=draft.proposed_stories,
        )
        primary_by_factor = {item.factor: item for item in primary_assessment.dimensions}
        reviewer_by_factor = {item.factor: item for item in reviewer_assessment.dimensions}
        arbitration_by_factor = {item.factor: item for item in arbitration}
        final_scorecard = []
        for item in build_scorecard(final_draft, story):
            primary_dimension = primary_by_factor[item.factor]
            reviewer_dimension = reviewer_by_factor[item.factor]
            resolution = arbitration_by_factor[item.factor]
            if resolution.selected_score == primary_dimension.score_most_likely:
                evidence_reason = primary_dimension.rationale
            elif resolution.selected_score == reviewer_dimension.score_most_likely:
                evidence_reason = reviewer_dimension.rationale
            else:
                evidence_reason = (
                    f"Primary evidence: {primary_dimension.rationale} "
                    f"Reviewer evidence: {reviewer_dimension.rationale}"
                )
            final_scorecard.append(
                item.model_copy(
                    update={
                        "reason": evidence_reason[:300],
                        "provenance": (
                            "model"
                            if primary_dimension.provenance == "model"
                            or reviewer_dimension.provenance == "model"
                            else "heuristic"
                        ),
                    }
                )
            )
        result = build_result(final_draft, story, manifest, final_scorecard)
        final_scores = {item["factor"]: int(item["score"]) for item in result["scorecard"]}
        audit = consistency_audit(
            primary_assessment,
            reviewer_assessment,
            disagreements,
            Calculation.model_validate(result["calculation"]),
            final_scores,
            story.stack,
            blind_review_executed,
        )
        # -- EAGLE governance package ------------------------------------------------------
        # Everything here is decided in code. §17 validates what is objectively checkable,
        # §20 is allowed to refuse to estimate, §10 anchors against history, §22 records what
        # would have to change for two runs to differ, and §29 says which layer failed.
        calculation_result = Calculation.model_validate(result["calculation"])
        validation = validate(final_scores, story.stack, board, calculation_result)
        gate = spike_gate(final_scores, story.stack)
        references = compare_references(
            story, final_scores, result["points"], reference_history or []
        )
        snapshot = build_snapshot(
            contract,
            canonical["input_hash"],
            story.stack,
            self.settings.model_id,
            len(proposals),
            len(reference_history or []),
        )
        failures = attribute_failure(validation, debate_outcome, board, len(proposals))
        if progress:
            progress(
                {
                    "stage": "eagle_validation",
                    "status": "completed" if validation.passed else "failed",
                    "label": (
                        f"{sum(i.passed for i in validation.rules)} of {len(validation.rules)} "
                        f"deterministic validation rules passed"
                    ),
                    "detail": (
                        "Objective rules are enforced in code, not in a prompt: factor count, "
                        "score range, evidence for elevated scores, and that the adjustments "
                        "still reconcile to the adjusted score."
                    ),
                    "evidence": {
                        "failed_rules": [i.rule for i in validation.failures()],
                        "spike_gate": gate.decision,
                        "spike_triggers": gate.triggered,
                    },
                }
            )
            progress(
                {
                    "stage": "eagle_reference",
                    "status": "completed",
                    "label": (
                        f"Closest of {len(reference_history or [])} historical estimate(s): "
                        f"{references.closest.points} points at "
                        f"{references.closest.similarity:.0%} similarity"
                        if references.closest
                        else "No comparable historical estimate to anchor against"
                    ),
                    "detail": references.note,
                    "evidence": {
                        "matches": [
                            {"title": m.title, "points": m.points, "similarity": m.similarity}
                            for m in references.matches
                        ],
                        "implied_range": references.implied_range,
                        "relative": references.relative_assessment,
                    },
                }
            )
        result["eagle"] = {
            "version": EAGLE_VERSION,
            "contract": contract.model_dump(mode="json"),
            "blackboard": {
                "records": [item.model_dump() for item in board.records],
                "sources": board.sources(),
            },
            "factor_aggregates": [item.model_dump() for item in aggregates],
            "findings": [item.model_dump() for item in eagle_findings],
            "debate": debate_outcome.model_dump(),
            "validation": validation.model_dump(),
            "spike_gate": gate.model_dump(),
            "references": references.model_dump(),
            "snapshot": snapshot.model_dump(mode="json"),
            "failure_attribution": [item.model_dump() for item in failures],
        }
        result["agentic_pipeline"] = {
            "version": PIPELINE_VERSION,
            "eagle_version": EAGLE_VERSION,
            "mode": pipeline_mode,
            "status": "HUMAN_REVIEW",
            "canonical_story": canonical,
            "readiness": readiness.model_dump(),
            "specialist_routes": [item.model_dump() for item in specialist_routes],
            "specialist_findings": [item.model_dump() for item in specialist_findings],
            "primary": primary_assessment.model_dump(),
            "reviewer": reviewer_assessment.model_dump(),
            "disagreements": [item.model_dump() for item in disagreements],
            "critic_challenges": [item.model_dump() for item in challenges],
            "arbitration": [item.model_dump() for item in arbitration],
            "consistency_audit": audit,
            "final_report": {
                "recommended_points": result["points"],
                "recommendation": result["recommendation"],
                "confidence": result["confidence"],
                "why_selected": result["plain_language_why"],
                "why_not_lower": result["detailed_reasoning"]["band_sensitivity"][
                    "explanation"
                ],
                "why_not_higher": (
                    "The next Fibonacci value requires additional scored evidence or a "
                    "framework gate; unsupported risk is not added to the estimate."
                ),
                "assumptions": result["assumptions"],
                "human_authority": (
                    "This is a decision-support recommendation. The delivery team owns "
                    "the final estimate and may accept, override, spike, or decompose it."
                ),
            },
            "human_review": {
                "required": True,
                "status": "pending",
                "options": ["accept", "override", "spike", "decompose"],
                "reason": (
                    "Every AI-assisted estimate requires an explicit human team decision."
                ),
            },
            "prompt_versions": {
                "primary": "estimate-primary-1.0",
                "reviewer": "estimate-blind-review-1.0",
            },
            "model_policy": {
                "model": self.settings.model_id,
                "serialized": True,
                # The blind review is conditional, so this must report what actually ran.
                # Stating 2 unconditionally described a second opinion that, for most
                # stories, never happened — in a record whose whole purpose is to be
                # checkable against the run.
                "independent_model_passes": 2 if blind_review_executed else 1,
                "blind_review_executed": blind_review_executed,
                "blind_review_reason": review_reason,
                "hidden_chain_of_thought_stored": False,
            },
        }

        if progress:
            provenance = result["evidence"]["scoring_provenance"]
            progress(
                {
                    "stage": "score_factors",
                    "status": "completed",
                    "label": f"16 factors scored ({provenance['model_scored']} by model)",
                    "evidence": provenance,
                }
            )
            calculation = result["calculation"]
            progress(
                {
                    "stage": "calculate",
                    "status": "completed",
                    "label": (
                        f"Base {calculation['base_sum']} "
                        f"+{calculation['base_adjustment_total']} base "
                        f"+{calculation['stack_adjustment_total']} stack "
                        f"= {calculation['adjusted_score']}"
                    ),
                    "detail": "Computed in application code from the scorecard, not by the model.",
                    "evidence": {
                        "adjusted_score": calculation["adjusted_score"],
                        "band": calculation["band"],
                        "points": calculation["points"],
                        "rules_fired": sum(
                            1 for step in calculation["steps"]
                            if step["applied"] and step["rule"] != "base_sum"
                        ),
                    },
                }
            )
            failed = [
                check["rule"] for check in result["evidence"]["policy_checks"]
                if not check["passed"]
            ]
            progress(
                {
                    "stage": "policy_gate",
                    "status": "completed" if not failed else "waiting",
                    "label": f"Decision: {result['recommendation'].replace('_', ' ')}",
                    "detail": result["recommendation_detail"],
                    "evidence": {
                        "gates_evaluated": len(result["evidence"]["policy_checks"]),
                        "gates_failed": failed or "none",
                        "confidence": result["confidence"],
                        "risk_flags": len(result["risk_flags"]),
                        "suggestions": len(result["suggestions"]),
                    },
                }
            )
            progress(
                {
                    "stage": "consistency_audit",
                    "status": "completed",
                    "label": f"Consistency audit: {audit['status'].replace('_', ' ').title()}",
                    "evidence": audit,
                }
            )
            progress(
                {
                    "stage": "human_review",
                    "status": "waiting",
                    "label": "Human team decision required",
                    "detail": result["agentic_pipeline"]["human_review"]["reason"],
                    "evidence": {
                        "options": result["agentic_pipeline"]["human_review"]["options"]
                    },
                }
            )
        return result


# --------------------------------------------------------------------------------------
# Upload and Jira ingestion (unchanged in behaviour; stack profile rides along)
# --------------------------------------------------------------------------------------

TARGET_ALIASES = {
    "title": ["title", "summary", "story title", "issue", "name"],
    "user_story": ["user story", "description", "story", "details", "requirement"],
    "acceptance_criteria": ["acceptance criteria", "acs", "ac", "criteria"],
    "technical_breakdown": ["technical breakdown", "technical notes", "implementation"],
    "existing_points": ["existing points", "story points", "points", "sp", "estimate"],
}


def _mapping(columns: list[str]) -> dict[str, str | None]:
    result: dict[str, str | None] = {}
    used: set[str] = set()
    for target, aliases in TARGET_ALIASES.items():
        scored: list[tuple[float, str]] = []
        for column in columns:
            if column in used:
                continue
            cleaned = re.sub(r"[^a-z0-9]+", " ", column.lower()).strip()
            score = max(
                1.0 if cleaned == alias else 0.9 if alias in cleaned else
                SequenceMatcher(None, cleaned, alias).ratio()
                for alias in aliases
            )
            scored.append((score, column))
        score, column = max(scored, default=(0.0, ""))
        result[target] = column if score >= 0.55 else None
        if result[target]:
            used.add(column)
    return result


def parse_upload(content: bytes, filename: str) -> dict[str, Any]:
    suffix = Path(filename).suffix.lower()
    if len(content) > 15 * 1024 * 1024:
        raise ValueError("File exceeds the 15 MB upload limit.")
    rows: list[dict[str, Any]]
    if suffix == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        rows = [dict(row) for row in csv.DictReader(io.StringIO(text))]
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Excel support is not installed. Run the setup script again.") from exc
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(values, [])]
        rows = [
            {headers[index]: "" if value is None else value for index, value in enumerate(row)}
            for row in values
        ]
    else:
        raise ValueError("Use a .csv or .xlsx file.")
    columns = list(rows[0].keys()) if rows else []
    return {
        "columns": columns,
        "suggested_mapping": _mapping(columns),
        "preview": rows[:20],
        "rows": rows[:100],
        "row_count": len(rows),
    }


def rows_to_stories(
    rows: list[dict[str, Any]],
    mapping: dict[str, str | None],
    stack: StackProfile | None = None,
) -> list[Story]:
    title_column = mapping.get("title")
    if not title_column:
        raise ValueError("Map a source column to Title before estimating.")
    profile = stack or StackProfile()
    stories: list[Story] = []
    for row in rows[:100]:
        title = str(row.get(title_column, "")).strip()
        if not title:
            continue
        raw_points = row.get(mapping.get("existing_points") or "", "")
        try:
            points = float(raw_points) if str(raw_points).strip() else None
        except (TypeError, ValueError):
            points = None
        stories.append(
            Story(
                title=title,
                user_story=str(row.get(mapping.get("user_story") or "", "")).strip(),
                acceptance_criteria=row.get(mapping.get("acceptance_criteria") or "", ""),
                technical_breakdown=(
                    str(row.get(mapping.get("technical_breakdown") or "", "")).strip() or None
                ),
                existing_points=points,
                source="upload",
                stack=profile,
            )
        )
    if not stories:
        raise ValueError("No valid story rows remain after mapping.")
    return stories


async def jira_issues(settings: Settings, project: str, query: str = "") -> list[dict]:
    if not (settings.jira_base_url and settings.jira_email and settings.jira_api_token):
        raise ValueError("Jira is not configured in the Devvy environment.")
    jql = f'project = "{project.replace(chr(34), "")}"'
    if query.strip():
        jql += f' AND text ~ "{query.replace(chr(34), "")}"'
    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.get(
            f"{settings.jira_base_url.rstrip('/')}/rest/api/3/search",
            params={
                "jql": jql,
                "maxResults": 100,
                "fields": (
                    "summary,description,status,labels,components,"
                    f"{settings.jira_story_points_field}"
                ),
            },
            auth=(settings.jira_email, settings.jira_api_token),
        )
        response.raise_for_status()
    issues = []
    for issue in response.json().get("issues", []):
        fields = issue.get("fields", {})
        issues.append(
            Story(
                title=fields.get("summary") or issue["key"],
                user_story=json.dumps(fields.get("description") or ""),
                existing_points=fields.get(settings.jira_story_points_field),
                key=issue["key"],
                status=(fields.get("status") or {}).get("name"),
                labels=fields.get("labels") or [],
                components=[item.get("name", "") for item in fields.get("components") or []],
                source="jira",
            ).model_dump()
        )
    return issues


async def write_jira_points(settings: Settings, issue_key: str, points: int) -> None:
    if not settings.jira_write_enabled:
        raise ValueError("Jira write-back is disabled by configuration.")
    if not (settings.jira_base_url and settings.jira_email and settings.jira_api_token):
        raise ValueError("Jira is not configured in the Devvy environment.")
    if not re.fullmatch(r"[A-Za-z][A-Za-z0-9_]+-\d+", issue_key):
        raise ValueError("Invalid Jira issue key.")
    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.put(
            f"{settings.jira_base_url.rstrip('/')}/rest/api/3/issue/{issue_key}",
            json={"fields": {settings.jira_story_points_field: points}},
            auth=(settings.jira_email, settings.jira_api_token),
        )
        response.raise_for_status()
